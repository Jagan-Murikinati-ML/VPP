#!/usr/bin/env python3
"""
Generate meterIdSiteIdTable for KQL from Excel file using openpyxl only.
Usage: python3 generate_meter_mapping.py [excel_filename]
"""

import sys
from openpyxl import load_workbook

# Accept filename as argument or use default
excel_file = sys.argv[1] if len(sys.argv) > 1 else 'Sites need both May and June interval data.xlsx'

try:
    wb = load_workbook(excel_file)
    ws = wb.active
    print(f"✓ Successfully read {excel_file}")
    print(f"✓ Sheet: {ws.title}\n")
except Exception as e:
    print(f"✗ Error reading file: {e}")
    exit(1)

# Get all rows
all_rows = list(ws.iter_rows(values_only=True))

# First row is headers
if not all_rows:
    print("✗ No data in file")
    exit(1)

headers = all_rows[0]
print("=" * 80)
print("HEADERS found:")
print(headers)
print()

# Find column indices (case-insensitive)
siteid_idx = None
meterid_idx = None

for i, header in enumerate(headers):
    if header is None:
        continue
    header_lower = str(header).lower().strip()
    if 'siteid' in header_lower or 'site_id' in header_lower or 'site' in header_lower.replace('_', ' '):
        siteid_idx = i
    if 'meterid' in header_lower or 'meter_id' in header_lower or 'leapmetid' in header_lower or 'leap' in header_lower:
        meterid_idx = i

if siteid_idx is None or meterid_idx is None:
    print(f"✗ Could not identify columns")
    print(f"Column indices found: siteid={siteid_idx}, meterid={meterid_idx}")
    print(f"Available columns: {list(enumerate(headers))}")
    exit(1)

print(f"✓ Column indices: siteId={siteid_idx}, meterId={meterid_idx}\n")

# Extract data rows (skip header)
pairs = []
for row in all_rows[1:]:
    if row[siteid_idx] is None or row[meterid_idx] is None:
        continue
    site_id = str(row[siteid_idx]).strip()
    meter_id = str(row[meterid_idx]).strip()
    if site_id and meter_id:
        pairs.append((site_id, meter_id))

print(f"✓ Extracted {len(pairs)} valid mappings\n")

if len(pairs) == 0:
    print("✗ No valid mappings found")
    exit(1)

# Display preview
print("=" * 80)
print("PREVIEW - First 5 mappings:")
print("=" * 80)
for i, (site, meter) in enumerate(pairs[:5]):
    print(f"{i+1}. siteId='{site}' -> meterId='{meter}'")
print()

# Generate KQL format
print("=" * 80)
print("KQL OUTPUT:")
print("=" * 80)

kql_lines = ["let meterIdSiteIdTable = datatable(siteId: string, meterId: string)"]
kql_lines.append("[")

# Format with line breaks for readability
current_line = "    "
line_items = []

for i, (site, meter) in enumerate(pairs):
    item = f"'{site}','{meter}'"
    line_items.append(item)

# Join with commas and wrap lines
for i, item in enumerate(line_items):
    if i > 0:
        current_line += ","
    current_line += item
    
    # Break line every ~3 items or at 100 chars
    if (i + 1) % 3 == 0 or len(current_line) > 100:
        kql_lines.append(current_line)
        current_line = "    "

# Add any remaining items
if current_line.strip() != "":
    kql_lines.append(current_line)

kql_lines.append("];")

kql_output = "\n".join(kql_lines)
print(kql_output)
print()

# Save to file
base_name = excel_file.replace('.xlsx', '').replace(' ', '_')
output_file = f"{base_name}_meterIdSiteIdTable.txt"
with open(output_file, 'w') as f:
    f.write(kql_output)

print("=" * 80)
print(f"✓ Saved to: {output_file}")
print(f"✓ Total mappings: {len(pairs)}")
print("=" * 80)
